# -*- coding: utf-8 -*-
"""
@Time ： 2023/5/7 15:29
@Auth ： 胡英俊(请叫我英俊)
@File ：Excel.py
@IDE ：PyCharm
@Motto：Never stop learning
"""
from openpyxl import *
import os
import time
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill


class Excel(object):
    # 构造函数，生成了一个self.wb，存储了excel的文件对象
    def __init__(self, excel_file_path=None):
        if excel_file_path and not os.path.exists(excel_file_path):
            print("设定的 excel 文件路径 %s 不存在" % excel_file_path)
            self.wb = None
            return
        if os.path.exists(excel_file_path):
            self.wb = load_workbook(excel_file_path)
            self.excel_file_path = excel_file_path

    # 加载某个路径的excel文件
    def load_excel_file(self, excel_file_path):
        if excel_file_path and not os.path.exists(excel_file_path):
            print("设定的 excel 文件路径 %s 不存在" % excel_file_path)
            self.wb = None
            return
        if os.path.exists(excel_file_path):
            self.wb = load_workbook(excel_file_path)
            self.excel_file_path = excel_file_path

    # 通过名字，设定你当前要操作的sheet名称
    def set_sheet_by_name(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
            return self.sheet
        else:
            print("设定的 sheet 名称 \"%s\" 不存在" % sheet_name)
            return None

    # 通过序号，设定你对当前要操作的名称
    def set_sheet_by_index(self, index):
        if index >= 0 and index <= len(self.wb.sheetnames):
            self.sheet = self.wb[self.wb.sheetnames[index - 1]]
            return self.sheet
        else:
            print("当前excel文件有%s个sheet，你设定的序号 %s 不在有效范围内" % (len(self.wb.sheetnames), index))
            return None

    # 创建一个新的sheet
    def create_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)
            print("创建 %s sheet成功！" % sheet_name)
        else:
            print(" %s sheet已经存在，无需创建！" % sheet_name)

    # 获得某个sheet中的所有行对象
    def get_all_rows(self):
        if self.sheet is None:
            print("没有设定要操作的 sheet!")
            return
        rows = []  # 存储一下行对象
        for row in self.sheet.rows:  # 遍历所有的行
            rows.append(row)

        return rows

    # 获取某个sheet中的所有行的值，返回一个二维列表
    def get_all_rows_values(self):
        if self.sheet is None:
            print("没有设定要操作的 sheet!")
            return
        rows = []  # 存储一下行对象
        for row in self.sheet.rows:  # 遍历所有的行
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            rows.append(tmp)

        return rows

    # 获取某个行对象
    def get_row(self, row_num):
        all_rows = self.get_all_rows()
        if all_rows:
            if row_num > 0 and row_num <= len(all_rows):
                return all_rows[row_num - 1]
            else:
                print("设定的行号 %s 不是有效范围，有效范围是1-%s" % (row_num, len(all_rows)))
        return None

    # 获取某个行中所有单元格的值，返回的是一维列表
    def get_row_values(self, row_num):
        row_values = []
        if self.get_row(row_num):
            for cell in self.get_row(row_num):
                row_values.append(cell.value)

        return row_values

    # 获取某个列对象
    def get_col(self, col_num):
        all_rows = self.get_all_rows()
        if all_rows:
            if col_num > 0 and col_num <= len(all_rows[0]):
                col = []
                for row in all_rows:
                    col.append(row[col_num - 1])
                return col
            else:
                print("设定的列号 %s 不是有效范围，有效范围是1-%s" % (col_num, len(all_rows[0])))
        return None

    # 获取某个列的所有单元格的值
    def get_col_values(self, col_num):
        col = self.get_col(col_num)
        if col:
            col_value = []
            for cell in col:
                col_value.append(cell.value)
            return col_value
        return None

    # 获取某个单元格对象，坐标从1开始
    def get_cell(self, row_num, col_num):
        all_rows = self.get_all_rows()
        try:
            if row_num <= 0 or col_num <= 0:
                raise Exception()
            return all_rows[row_num - 1][col_num - 1]
        except:
            print("设置的单元格坐标:行号%s 列号%s有误，无法读取单元格" % (row_num, col_num))
            return None

    # 获取某个单元格的值，坐标从1开始
    def get_cell_value(self, row_num, col_num):
        if self.get_cell(row_num, col_num):
            return self.get_cell(row_num, col_num).value
        return None

    # 写入某一行，row_data是一个一维列表，作为行数据来写入
    def write_line(self, row_data, row_no=None, background_color=None):
        row_data = [str(i) for i in row_data]
        for data_id in range(len(row_data)):
            if row_data[data_id] == 'None':
                row_data[data_id] = "-"
        print(row_data)
        bd = Border(left=Side(border_style="thin",
                              color='FF001000'),
                    right=Side(border_style="thin",
                               color='FF110000'),
                    top=Side(border_style="thin",
                             color='FF110000'),
                    bottom=Side(border_style="thin",
                                color='FF110000')
                    )

        if not isinstance(row_data, (list, tuple)):
            print("写入行的数据不是列表或元组数据，无法写入！")
            return None

        if not row_no:
            # 获得最大行号
            row_num = self.sheet.max_row + 1
            print(row_num)
            for col_num in range(len(row_data)):
                self.sheet.cell(row=row_num, column=col_num + 1, value=row_data[col_num])
                self.get_cell(row_num, col_num + 1).border = bd
                if background_color:
                    self.get_cell(row_num, col_num + 1).fill = PatternFill("solid", fgColor=background_color)
                if row_data[col_num] == "失败" or "fail" in row_data[col_num] or "失" in row_data[col_num] or "败" in \
                        row_data[col_num]:
                    self.get_cell(row_num, col_num + 1).fill = PatternFill("solid", fgColor="FF0000")
                self.save()
        else:
            print("指定的行号：%s" % row_no)
            for col_num in range(len(row_data)):
                self.sheet.cell(row=row_no, column=col_num + 1, value=row_data[col_num])
                self.get_cell(row_no, col_num + 1).border = bd
        self.save()

    def write_cell(self, row_no, col_no, value, background_colour=None):
        bd = Border(left=Side(border_style="thin",
                              color='FF001000'),
                    right=Side(border_style="thin",
                               color='FF110000'),
                    top=Side(border_style="thin",
                             color='FF110000'),
                    bottom=Side(border_style="thin",
                                color='FF110000')
                    )
        if row_no <= 0 or col_no <= 0:
            print("设定的坐标必须大于等于1开始")
            return

        if not isinstance(row_no, int):
            print("设定的行号必须是数字，你设定的是%s" % row_no)

        if not isinstance(col_no, int):
            print("设定的列号必须是数字，你设定的是%s" % col_no)

        if value:
            self.sheet.cell(row=row_no, column=col_no, value=value)
            if background_colour:
                self.sheet.cell(row=row_no, column=col_no).fill = PatternFill("solid", fgColor=background_colour)
            if value == "失败" or "fail" in value or "失" in value or "败" in value:
                self.get_cell(row_no, col_no).fill = PatternFill("solid", fgColor="FF0000")
        self.save()

    def write_cell_current_time(self, row_no, col_no):
        date = str(time.localtime().tm_year) + "年" + str(time.localtime().tm_mon) + "月" + str(
            time.localtime().tm_mday) + "日"
        now_time = str(time.localtime().tm_hour) + "时" + str(time.localtime().tm_min) + "分" + str(
            time.localtime().tm_sec) + "秒"
        now = date + " " + now_time
        self.write_cell(row_no, col_no, now)

    # 获取当前操作的excel对象的绝对路径
    def get_excel_file_path(self):
        return self.excel_file_path

    # 保存当前excel文件对象所做的所有修改
    def save(self):
        self.wb.save(self.get_excel_file_path())


if __name__ == "__main__":
    wb = Excel("e:\\testexcel.xlsx")
    sheet = wb.set_sheet_by_name("Sheet1")

    '''
    wb.write_line([1,2,3],1)
    wb.write_line([4,5,6],2)
    wb.write_line(["你","我","他"],background_color="FF0000")
    wb.write_line(["今天","昨天","明天"],background_color="008000")
    wb.write_line(["今天", "昨天", "明天"],background_color="DDEEDD")
    wb.write_line(["失败", "败", "好"], background_color="DDEEDD")
    #FF0000:红色
    #008000：绿色
    '''
    wb.write_cell(15, 6, "今天天气不错！")
    wb.write_cell(15, 7, "成功", background_colour="008000")
    wb.write_cell(15, 8, "失败！")
    wb.write_cell_current_time(15, 9)
    '''
        print(sheet)
        sheet = wb.set_sheet_by_index(1)
        print(sheet)

        sheet = wb.set_sheet_by_index(2)
        print(sheet)

        #sheet = wb.set_sheet_by_index(3)
        #print(sheet)

        rows = wb.get_all_rows()
        print(rows)

        rows_values = wb.get_all_rows_values()
        print(rows_values )

        row = wb.get_row(2)
        print(row)

        row_values = wb.get_row_values(2)
        print(row_values)

        col = wb.get_col(2)
        print(col)

        col_values = wb.get_col_values(100)
        print(col_values)

        cell = wb.get_cell(2,2)
        print(cell)

        cell_value = wb.get_cell_value(2,2)
        print(cell_value)

        wb.create_sheet("测试数据")
        wb.save()
        '''
